#!/usr/bin/env python
# -*- coding: utf-8 -*-
import psycopg2
import datetime
import psycopg2.extras
import logging
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from entity import get_connection

def get_vw_rep_form_12(conn):
    curr = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    curr.execute("SELECT * FROM vw_rep_form_12")
    data = curr.fetchall()
    #['р-ны Новосибирска', '00.01', 'Тестовая организация', 'Текстовая запись ОКПДТР.1', '0.1', '7211', 'СПО', 'Тестовая запись1', 1, 3, 1, 985, 1, 10, 11, 12, 13, 14, 15]
    logging.debug(data)
    logging.debug(data[0]['raion'])

    len_data = len(data)
    for i in range(0, len_data):
        n6 = ((int(data[i]['y_0'])) + (int(data[i]['y_1'])) + (int(data[i]['y_2'])) + (int(data[i]['y_3'])) + (int(data[i]['y_4'])) + (int(data[i]['y_5'])))
        data[i]['y_6'] = n6
    return data

def render_vw_rep_form_12(conn, templ_path):
    data = get_vw_rep_form_12(conn)
    len_data = len(data)
    sum_1 = 0
    sum_2 = 0
    sum_3 = 0
    sum_4 = 0
    sum_5 = 0
    sum_6 = 0
    for i in range(0, len_data):
        sum_1 = sum_1 + int(data[i]['y_0'])
        sum_2 = sum_2 + int(data[i]['y_1'])
        sum_3 = sum_3 + int(data[i]['y_2'])
        sum_4 = sum_4 + int(data[i]['y_3'])
        sum_5 = sum_5 + int(data[i]['y_4'])
        sum_6 = sum_6 + int(data[i]['y_5'])
    data_sum = [sum_1, sum_2, sum_3, sum_4, sum_5, sum_6]
    #открываем файл
    tpl_name = templ_path+'templates/{0}.xlsx'.format('rep_form_12_ex')
    wb = load_workbook(tpl_name)
    ws = wb.active
    side = Side(border_style='medium', color="FF000000")
    font = Font(name='Times New Roman', size=10)
    border = Border(left=side,
                    right=side,
                    top=side,
                    bottom=side)

    for i in range(0, len_data):
        ws['A{0}'.format(i+4)] = str(data[i]['raion'])
        ws['A{0}'.format(i+4)].border = border
        ws['A{0}'.format(i+4)].font = font
        ws['B{0}'.format(i+4)] = str(data[i]['okved_code'])
        ws['B{0}'.format(i+4)].border = border
        ws['B{0}'.format(i+4)].font = font
        ws['C{0}'.format(i+4)] = str(data[i]['organization_name'])
        ws['C{0}'.format(i+4)].border = border
        ws['C{0}'.format(i+4)].font = font
        ws['D{0}'.format(i+4)] = str(data[i]['okpdtr_name'])
        ws['D{0}'.format(i+4)].border = border
        ws['D{0}'.format(i+4)].font = font
        ws['E{0}'.format(i+4)] = str(data[i]['okpdtr_code'])
        ws['E{0}'.format(i+4)].border = border
        ws['E{0}'.format(i+4)].font = font
        ws['F{0}'.format(i+4)] = str(data[i]['okz_code'])
        ws['F{0}'.format(i+4)].border = border
        ws['F{0}'.format(i+4)].font = font
        ws['G{0}'.format(i+4)] = str(data[i]['case'])
        ws['G{0}'.format(i+4)].border = border
        ws['G{0}'.format(i+4)].font = font
        ws['H{0}'.format(i+4)] = str(data[i]['okved_name'])
        ws['H{0}'.format(i+4)].border = border
        ws['H{0}'.format(i+4)].font = font
        ws['I{0}'.format(i+4)] = str(data[i]['y_0'])
        ws['I{0}'.format(i+4)].border = border
        ws['I{0}'.format(i+4)].font = font
        ws['J{0}'.format(i+4)] = str(data[i]['y_1'])
        ws['J{0}'.format(i+4)].border = border
        ws['J{0}'.format(i+4)].font = font
        ws['K{0}'.format(i+4)] = str(data[i]['y_2'])
        ws['K{0}'.format(i+4)].border = border
        ws['K{0}'.format(i+4)].font = font
        ws['L{0}'.format(i+4)] = str(data[i]['y_3'])
        ws['L{0}'.format(i+4)].border = border
        ws['L{0}'.format(i+4)].font = font
        ws['M{0}'.format(i+4)] = str(data[i]['y_4'])
        ws['M{0}'.format(i+4)].border = border
        ws['M{0}'.format(i+4)].font = font
        ws['N{0}'.format(i+4)] = str(data[i]['y_5'])
        ws['N{0}'.format(i+4)].border = border
        ws['N{0}'.format(i+4)].font = font
        ws['O{0}'.format(i+4)] = str(data[i]['y_6'])
        ws['O{0}'.format(i+4)].border = border
        ws['O{0}'.format(i+4)].font = font

    ws['A{0}'.format(i+4+1)] ="Итого:"
    ws['I{0}'.format(i+4+1)] = str(data_sum[0])
    ws['J{0}'.format(i+4+1)] = str(data_sum[1])
    ws['K{0}'.format(i+4+1)] = str(data_sum[2])
    ws['L{0}'.format(i+4+1)] = str(data_sum[3])
    ws['M{0}'.format(i+4+1)] = str(data_sum[4])
    ws['N{0}'.format(i+4+1)] = str(data_sum[5])

    ws['A{0}'.format(i+4+1)].border = border
    ws['B{0}'.format(i+4+1)].border = border
    ws['C{0}'.format(i+4+1)].border = border
    ws['D{0}'.format(i+4+1)].border = border
    ws['E{0}'.format(i+4+1)].border = border
    ws['F{0}'.format(i+4+1)].border = border
    ws['G{0}'.format(i+4+1)].border = border
    ws['H{0}'.format(i+4+1)].border = border
    ws['I{0}'.format(i+4+1)].border = border
    ws['J{0}'.format(i+4+1)].border = border
    ws['K{0}'.format(i+4+1)].border = border
    ws['L{0}'.format(i+4+1)].border = border
    ws['M{0}'.format(i+4+1)].border = border
    ws['N{0}'.format(i+4+1)].border = border
    ws['O{0}'.format(i+4+1)].border = border

    ws['A{0}'.format(i+4+1)].font = font
    ws['B{0}'.format(i+4+1)].font = font
    ws['C{0}'.format(i+4+1)].font = font
    ws['D{0}'.format(i+4+1)].font = font
    ws['E{0}'.format(i+4+1)].font = font
    ws['F{0}'.format(i+4+1)].font = font
    ws['G{0}'.format(i+4+1)].font = font
    ws['H{0}'.format(i+4+1)].font = font
    ws['I{0}'.format(i+4+1)].font = font
    ws['J{0}'.format(i+4+1)].font = font
    ws['K{0}'.format(i+4+1)].font = font
    ws['L{0}'.format(i+4+1)].font = font
    ws['M{0}'.format(i+4+1)].font = font
    ws['N{0}'.format(i+4+1)].font = font
    ws['O{0}'.format(i+4+1)].font = font

    ws['A{0}'.format(i+4 +3)]= "Первая цифра кода ОКЗ указывает на уровень образования: 1-2 специалисты высшего уровня квалификации (ВО); 3-8 – специалисты среднего уровня квалификации и и квалифицированные рабочие (СПО); 9 – неквалифицированные рабочие (НР)"
    ws.merge_cells(start_row=(i+4+3), start_column=1, end_row=(i+4+3), end_column=15)
    today = datetime.datetime.today()
    date_here = (today.strftime("%Y.%m.%d-%H.%M.%S"))
    logging.debug('xlsx generated')
    #сохраняем
    wb.save(templ_path + 'generated/rep_form_12-{0}.xlsx'.format(date_here))


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG, format='%(lineno)d %(asctime)s %(message)s')
    res = render_vw_rep_form_12(get_connection(), "")
    logging.debug(res)