#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Загружаем данные проф. стандартов из Excel-файлов"""
import logging
import glob
import psycopg2
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from iastables import ps_classes


MAX_ROWS = 100
START_COLUMN = 1
PARAM_COLS = 4
START_ROW = 4
OKZ_COLUMN = 5
OKVED_COLUMN = 6
OKSO_COLUMN = 7
TF_CNT_COLUMN = 8
COLS_IN_TF = 8
TF_LEVEL = 1
TF_PROFS = 2
TF_REQS_STAGE = 3
TF_REQS_OKSO = 4
TF_REQS_SPECIAL = 5
TF_OKZ = 6
TF_OKDPTR = 7
TF_OKSO = 8


def cols_generator(col, cnt):
    """
    генератор индексов для столбцов аспекта всех трудовых функций
    :param col: конкретный столбей аспекта ТФ
    :param cnt: кол-во трудовых функций
    :return: генератор
    """
    return (TF_CNT_COLUMN + COLS_IN_TF * io + col for io in range(cnt))


def get_row_data(sheet, vrange):
    """
    данные ячеек в строке в диапазоне
    :param sheet: лист
    :param vrange: диапазон столбцов
    :return: список знаячений
    """
    res = [sheet.cell(row=START_ROW, column=j).value for j in vrange]
    logging.debug(res)
    return res


def get_column_data(sheet, col):
    """
    данные ячеек в столбце до первой пустой
    :param sheet: лист
    :param col: столбец
    :return: список значений
    """
    res = [sheet.cell(row=io, column=col).value for io in range(START_ROW, MAX_ROWS)
           if not sheet.cell(row=io, column=col).value is None]
    logging.debug(res)
    return res


def get_cell_data(sheet, vrow, vcol, vtype=int):
    """
    данные конкретныой ячейки, приведенные к типу
    :param sheet: лист
    :param vrow: строка
    :param vcol: столбец
    :param vtype: тип
    :return: значений ячейки
    """
    res = vtype(sheet.cell(row=vrow, column=vcol).value)
    logging.debug(res)
    return res


def get_tf_data_one(sheet, col, cnt):
    """
    данные конкретного аспекта трудовых функций
    :param sheet: лист
    :param col: столбец аспекта
    :param cnt: кол-во функций
    :return: список значений
    """
    res = [sheet.cell(row=START_ROW, column=io).value for io in cols_generator(col, cnt)]
    logging.debug(res)
    return res


def get_tf_data_all(sheet, col, cnt):
    """
    данные конкретного аспекта трудовых функций
    :param sheet: лист
    :param col: столбец аспекта
    :param cnt: кол-во трудовых функций
    :return: список списов значений
    """
    res = [[sheet.cell(row=io, column=c).value
            for io in range(START_ROW, MAX_ROWS)
            if not sheet.cell(row=io, column=c).value is None]
           for c in cols_generator(col, cnt)]
    logging.debug(res)
    return res


def process_file(workbook, vdb):
    """
    читаем все листы из файла
    :param workbook: книга
    :param vdb: соединение с б.д.
    :return:
    """
    for sheet in workbook:
        logging.debug(sheet)
        process_sheet(sheet, vdb)


def process_sheet(sheet, vdb):
    """
    читаем данные из листа
    :param sheet: лист
    :param vdb: соединение с б.д.
    :return:
    """
    if sheet.cell(row=START_ROW, column=1).value is None:
        return
    ps_params = get_row_data(sheet, range(START_COLUMN, PARAM_COLS+1))
    ps = ps_classes.get_ps_standart(vdb, sheet.title)
    ps.name = ps_params[2]
    ps.date_accepted = ps_params[3]
    ps.remark = ps_params[0]
    ps_okz = get_column_data(sheet, OKZ_COLUMN)
    ps_okved = get_column_data(sheet, OKVED_COLUMN)
    ps_okso = get_column_data(sheet, OKSO_COLUMN)
    cnt_tf = get_cell_data(sheet, START_ROW, TF_CNT_COLUMN)
    ps.tf_cnt = cnt_tf
    vdb.add(ps)
    # logging.debug(ps.id)
    ps_okz_lines = [ps_classes.get_ps_okz(vdb, ps, x) for x in ps_okz]
    for t in ps_okz_lines:
        try:
            vdb.add(t)
            vdb.commit()
        except Exception as e:
            logging.error(e)
            vdb.rollback()
    ps_okso_lines = [ps_classes.get_ps_okso(vdb, ps, x) for x in ps_okso]
    for t in ps_okso_lines:
        try:
            vdb.add(t)
            vdb.commit()
        except Exception as e:
            logging.error(e)
            vdb.rollback()
    ps_okved_lines = [ps_classes.get_ps_okved(vdb, ps, x) for x in ps_okved]
    for t in ps_okved_lines:
        try:
            vdb.add(t)
            vdb.commit()
        except Exception as e:
            logging.error(e)
            vdb.rollback()
    tf_levels = get_tf_data_one(sheet, TF_LEVEL, cnt_tf)
    ps_tfs = [ps_classes.get_ps_tf(vdb, ps, xi, x) for xi, x in enumerate(tf_levels)]
    for t in ps_tfs:
        try:
            vdb.add(t)
            vdb.commit()
        except Exception as e:
            logging.error(e)
            vdb.rollback()
    tf_profs = get_tf_data_all(sheet, TF_PROFS, cnt_tf)
    tf_stages = get_tf_data_all(sheet, TF_REQS_STAGE, cnt_tf)
    tf_oksos = get_tf_data_all(sheet, TF_REQS_OKSO, cnt_tf)
    # tf_special = get_tf_data_all(sheet, TF_REQS_SPECIAL, cnt_tf)
    tf_okz = get_tf_data_all(sheet, TF_OKZ, cnt_tf)
    tf_okdptr = get_tf_data_all(sheet, TF_OKDPTR, cnt_tf)
    tf_okso = get_tf_data_all(sheet, TF_OKSO, cnt_tf)
    for it, t in enumerate(ps_tfs):
        logging.debug('num - {} val - {} tf_profs -{} cnt_tf - {} '.format(it, t, tf_profs, cnt_tf))
        for tp in tf_profs[it]:
            try:
                a = ps_classes.get_ps_tf_prof(vdb, t, tp)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
        for ts in tf_stages[it]:
            try:
                a = ps_classes.get_ps_tf_stage(vdb, t, ts)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
        for too in tf_oksos[it]:
            try:
                a = ps_classes.get_ps_educ_reqs(vdb, t, too)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
        for to1 in tf_okso[it]:
            try:
                a = ps_classes.get_ps_tf_okso(vdb, t, to1)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
        for tsp in tf_okz[it]:
            try:
                a = ps_classes.get_ps_access_reqs(vdb, t, tsp)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
        for tsr in tf_okdptr[it]:
            try:
                a = ps_classes.get_ps_okdptr(vdb, t, tsr)
                vdb.add(a)
                vdb.commit()
            except Exception as e:
                logging.error(e)
                vdb.rollback()
    vdb.commit()


if __name__ == '__main__':
    logging.basicConfig(level=logging.ERROR, format='%(lineno)d %(asctime)s %(message)s')
    _db = psycopg2.connect("dbname='ias' user='dba' host='217.71.129.139' password='UdrqNjWx' port='4194'")
    logging.debug(_db)

    Session = sessionmaker(bind=create_engine('postgresql://dba:UdrqNjWx@217.71.129.139:4194/ias'))
    sess = Session()
    logging.debug(sess)
    for i, xls_path in enumerate(glob.glob("D:\\ias\\server\\data\\ps\\*.xlsx")):
        logging.info(xls_path)
        process_file(load_workbook(filename=xls_path), sess)
