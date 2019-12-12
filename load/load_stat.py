#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Про что этот файл?"""
import logging
from sqlalchemy import Column, DateTime, Integer, Numeric, String, text, Float
from sqlalchemy.ext.declarative import declarative_base
from entity import get_session
from openpyxl import load_workbook
import csv
import re

Base = declarative_base()
metadata = Base.metadata

COLS1 = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
         'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR',
         'A_S', 'AT', 'AU', 'AV', 'AW']


COLS2 = ['AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ',
         'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC']


COLS3 = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
         'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO']


COLS4 = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
         'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ',
         'AR', 'A_S', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE']


class LdStatDatum(Base):
    __tablename__ = 'ld_stat_data'

    id = Column(Integer, primary_key=True, server_default=text("nextval('ld_stat_data_id_seq'::regclass)"))
    region = Column(String(1000))
    okved = Column(String(1000))
    year_for = Column(Integer)
    e = Column(Numeric(12, 2))
    f = Column(Numeric(12, 2))
    g = Column(Numeric(12, 2))
    h = Column(Numeric(12, 2))
    i = Column(Numeric(12, 2))
    j = Column(Numeric(12, 2))
    k = Column(Numeric(12, 2))
    l = Column(Numeric(12, 2))
    m = Column(Numeric(12, 2))
    n = Column(Numeric(12, 2))
    o = Column(Numeric(12, 2))
    p = Column(Numeric(12, 2))
    q = Column(Numeric(12, 2))
    r = Column(Numeric(12, 2))
    s = Column(Numeric(12, 2))
    t = Column(Numeric(12, 2))
    u = Column(Numeric(12, 2))
    v = Column(Numeric(12, 2))
    w = Column(Numeric(12, 2))
    x = Column(Numeric(12, 2))
    y = Column(Numeric(12, 2))
    z = Column(Numeric(12, 2))
    aa = Column(Numeric(12, 2))
    ab = Column(Numeric(12, 2))
    ac = Column(Numeric(12, 2))
    ad = Column(Numeric(12, 2))
    ae = Column(Numeric(12, 2))
    af = Column(Numeric(12, 2))
    ag = Column(Numeric(12, 2))
    ah = Column(Numeric(12, 2))
    ai = Column(Numeric(12, 2))
    aj = Column(Numeric(12, 2))
    ak = Column(Numeric(12, 2))
    al = Column(Numeric(12, 2))
    am = Column(Numeric(12, 2))
    an = Column(Numeric(12, 2))
    ao = Column(Numeric(12, 2))
    ap = Column(Numeric(12, 2))
    aq = Column(Numeric(12, 2))
    ar = Column(Numeric(12, 2))
    a_s = Column(Numeric(12, 2))
    at = Column(Numeric(12, 2))
    au = Column(Numeric(12, 2))
    av = Column(Numeric(12, 2))
    aw = Column(Numeric(12, 2))
    ax = Column(Numeric(12, 2))
    ay = Column(Numeric(12, 2))
    az = Column(Numeric(12, 2))
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class LdStatData2(Base):
    __tablename__ = 'ld_stat_data2'

    id = Column(Integer, primary_key=True, server_default=text("nextval('ld_stat_data2_id_seq'::regclass)"))
    region = Column(String(1000))
    year_for = Column(Integer)
    az = Column(Numeric(12, 2))
    ba = Column(Numeric(12, 2))
    bb = Column(Numeric(12, 2))
    bc = Column(Numeric(12, 2))
    bd = Column(Numeric(12, 2))
    be = Column(Numeric(12, 2))
    bf = Column(Numeric(12, 2))
    bg = Column(Numeric(12, 2))
    bh = Column(Numeric(12, 2))
    bi = Column(Numeric(12, 2))
    bj = Column(Numeric(12, 2))
    bk = Column(Numeric(12, 2))
    bl = Column(Numeric(12, 2))
    bm = Column(Numeric(12, 2))
    bn = Column(Numeric(12, 2))
    bo = Column(Numeric(12, 2))
    bp = Column(Numeric(12, 2))
    bq = Column(Numeric(12, 2))
    br = Column(Numeric(12, 2))
    bs = Column(Numeric(12, 2))
    bt = Column(Numeric(12, 2))
    bu = Column(Numeric(12, 2))
    bv = Column(Numeric(12, 2))
    bw = Column(Numeric(12, 2))
    bx = Column(Numeric(12, 2))
    by = Column(Numeric(12, 2))
    bz = Column(Numeric(12, 2))
    ca = Column(Numeric(12, 2))
    cb = Column(Numeric(12, 2))
    cc = Column(Numeric(12, 2))
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class StatParamName2(Base):
    __tablename__ = 'stat_param_name2'

    id = Column(Integer, primary_key=True, server_default=text("nextval('stat_param_name_id_seq'::regclass)"))
    ord_num = Column(Integer, nullable=False, unique=True)
    col_name = Column(String(3))
    name = Column(String(512), nullable=False, unique=True)
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class LdStatData3(Base):
    __tablename__ = 'ld_stat_data3'

    id = Column(Integer, primary_key=True, server_default=text("nextval('ld_stat_data3_id_seq'::regclass)"))
    region = Column(String(1000))
    type_migration = Column(String(1000))
    year_for = Column(Integer)
    d = Column(Integer)
    e = Column(Integer)
    f = Column(Integer)
    g = Column(Integer)
    h = Column(Integer)
    i = Column(Integer)
    j = Column(Integer)
    k = Column(Integer)
    l = Column(Integer)
    m = Column(Integer)
    n = Column(Integer)
    o = Column(Integer)
    p = Column(Integer)
    q = Column(Integer)
    r = Column(Integer)
    s = Column(Integer)
    t = Column(Integer)
    u = Column(Integer)
    v = Column(Integer)
    w = Column(Integer)
    x = Column(Integer)
    y = Column(Integer)
    z = Column(Integer)
    aa = Column(Integer)
    ab = Column(Integer)
    ac = Column(Integer)
    ad = Column(Integer)
    ae = Column(Integer)
    af = Column(Integer)
    ag = Column(Integer)
    ah = Column(Integer)
    ai = Column(Integer)
    aj = Column(Integer)
    ak = Column(Integer)
    al = Column(Integer)
    am = Column(Integer)
    an = Column(Integer)
    ao = Column(Integer)
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class StatParamName3(Base):
    __tablename__ = 'stat_param_name3'

    id = Column(Integer, primary_key=True, server_default=text("nextval('stat_param_name3_id_seq'::regclass)"))
    ord_num = Column(Integer, nullable=False, unique=True)
    col_name = Column(String(3))
    name = Column(String(512), nullable=False, unique=True)
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class StatParamName4(Base):
    __tablename__ = 'stat_param_name4'

    id = Column(Integer, primary_key=True, server_default=text("nextval('stat_param_name4_id_seq'::regclass)"))
    ord_num = Column(Integer, nullable=False, unique=True)
    col_name = Column(String(3))
    name = Column(String(512), nullable=False, unique=True)
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class LdStatData4(Base):
    __tablename__ = 'ld_stat_data4'

    id = Column(Integer, primary_key=True, server_default=text("nextval('ld_stat_data4_id_seq'::regclass)"))
    region = Column(String(1000))
    type_age = Column(String(1000))
    year_for = Column(Integer)
    d = Column(Integer)
    e = Column(Integer)
    f = Column(Integer)
    g = Column(Integer)
    h = Column(Integer)
    i = Column(Integer)
    j = Column(Integer)
    k = Column(Integer)
    l = Column(Integer)
    m = Column(Integer)
    n = Column(Integer)
    o = Column(Integer)
    p = Column(Integer)
    q = Column(Integer)
    r = Column(Integer)
    s = Column(Integer)
    t = Column(Integer)
    u = Column(Integer)
    v = Column(Integer)
    w = Column(Integer)
    x = Column(Integer)
    y = Column(Integer)
    z = Column(Integer)
    aa = Column(Integer)
    ab = Column(Integer)
    ac = Column(Integer)
    ad = Column(Integer)
    ae = Column(Integer)
    af = Column(Integer)
    ag = Column(Integer)
    ah = Column(Integer)
    ai = Column(Integer)
    aj = Column(Integer)
    ak = Column(Integer)
    al = Column(Integer)
    am = Column(Integer)
    an = Column(Integer)
    ao = Column(Integer)
    ap = Column(Integer)
    aq = Column(Integer)
    ar = Column(Integer)
    a_s = Column(Integer)
    at = Column(Integer)
    au = Column(Integer)
    av = Column(Integer)
    aw = Column(Integer)
    ax = Column(Integer)
    ay = Column(Integer)
    az = Column(Integer)
    ba = Column(Integer)
    bb = Column(Integer)
    bc = Column(Integer)
    bd = Column(Integer)
    be = Column(Integer)
    dateadd = Column(DateTime, nullable=False, server_default=text("now()"))


class StatAgeDeath(Base):
    __tablename__ = 'stat_age_death'

    id = Column(Integer, primary_key=True, server_default=text("nextval('stat_age_death_id_seq'::regclass)"))
    age = Column(Integer, nullable=False)
    prob = Column(Float(53))
    date_add = Column(DateTime, nullable=False, server_default=text("now()"))


def char_range(c1, c2):
    """Generates the characters from `c1` to `c2`, inclusive."""
    for c in range(ord(c1), ord(c2) + 1):
        yield chr(c)


def process_file(wb, sess):
    sheet = wb[u"ОКВЭД"]
    logging.debug(sheet)
    # проходим со второй строки
    # до 1660
    i = 1
    #for j, c in enumerate(COLS1):
    #    stat_name = StatParamName(ord_num=j, col_name=c, name=sheet.cell(row=i, column=5+j).value)
    #    sess.add(stat_name)
    for i in range(2, 1661):
        val = sheet.cell(row=i, column=3).value
        if val is not None:
            data_line = {}
            try:
                okved_num = re.search(r"\d\d*", val).group(0)
                mrigo = sheet.cell(row=i, column=2).value
                year_for = sheet.cell(row=i, column=4).value
                data_line = {'region': mrigo, 'okved': okved_num, 'year_for': year_for}
                for j, c in enumerate(COLS1):
                    try:
                        data_line[c.lower()] = float(str(sheet.cell(row=i, column=5+j).value).replace(',', '.').replace(' ', ''))
                    except:
                        pass
                logging.debug(data_line)
            except KeyError:
                pass
            lsd = LdStatDatum(**data_line)
            sess.add(lsd)
    sess.commit()
    sess.close()


def process_file2(wb, sess):
    sheet = wb[u"ОКВЭД"]
    logging.debug(sheet)
    # проходим со второй строки
    # до 1660
    i = 1
    for j, c in enumerate(COLS2):
        stat_name = StatParamName2(ord_num=j, col_name=c, name=sheet.cell(row=i, column=52+j).value)
        sess.add(stat_name)
    for i in range(2, 1661):
        val = sheet.cell(row=i, column=52).value
        if val is not None:
            data_line = {}
            try:
                mrigo = sheet.cell(row=i, column=2).value
                year_for = sheet.cell(row=i, column=4).value
                data_line = {'region': mrigo, 'year_for': year_for}
                for j, c in enumerate(COLS2):
                    try:
                        data_line[c.lower()] = float(str(sheet.cell(row=i, column=52+j).value).replace(',', '.').replace(' ', ''))
                    except:
                        pass
                logging.debug(data_line)
            except KeyError:
                pass
            lsd = LdStatData2(**data_line)
            sess.add(lsd)
    sess.commit()
    sess.close()


def process_file3(wb, sess):
    sheet = wb[u"миграция "]
    logging.debug(sheet)
    # проходим со второй строки
    # до 1660
    hrs_cnt = 4
    hrs = [''] * hrs_cnt
    for j, c in enumerate(COLS3):
        for i in range(hrs_cnt):
            if sheet.cell(row=i+1, column=4+j).value is not None:
                hrs[i:] = [''] * len(hrs[i:])
                hrs[i] = sheet.cell(row=i+1, column=4+j).value
        stat_name = StatParamName3(ord_num=j, col_name=c, name='|'.join(hrs))
        sess.add(stat_name)
    for i in range(5, 654):
        val = sheet.cell(row=i, column=3).value
        if val is not None:
            data_line = {}
            try:
                mrigo = sheet.cell(row=i, column=1).value
                type_migration = sheet.cell(row=i, column=2).value
                year_for = sheet.cell(row=i, column=3).value
                data_line = {'region': mrigo, 'year_for': year_for, 'type_migration': type_migration}
                for j, c in enumerate(COLS3):
                    try:
                        data_line[c.lower()] = int(sheet.cell(row=i, column=4+j).value)
                    except:
                        pass
                logging.debug(data_line)
            except KeyError:
                pass
            lsd = LdStatData3(**data_line)
            sess.add(lsd)
    sess.commit()
    sess.close()


def process_file4(wb, sess):
    sheet = wb[u"Демография "]
    logging.debug(sheet)
    # проходим со второй строки
    # до 1660
    hrs_cnt = 5
    hrs = [''] * hrs_cnt
    for j, c in enumerate(COLS4):
        for i in range(hrs_cnt):
            if sheet.cell(row=i+1, column=4+j).value is not None:
                hrs[i:] = [''] * len(hrs[i:])
                hrs[i] = sheet.cell(row=i+1, column=4+j).value
        logging.debug(c)
        logging.debug('|'.join(hrs))
        stat_name = StatParamName4(ord_num=j, col_name=c, name='|'.join(hrs))
        sess.add(stat_name)
    for i in range(6, 4541):
        val = sheet.cell(row=i, column=4).value
        if val is not None:
            data_line = {}
            try:
                mrigo = sheet.cell(row=i, column=1).value
                type_age = sheet.cell(row=i, column=3).value
                year_for = sheet.cell(row=i, column=2).value
                data_line = {'region': mrigo, 'year_for': year_for, 'type_age': type_age}
                for j, c in enumerate(COLS4):
                    try:
                        data_line[c.lower()] = int(sheet.cell(row=i, column=4+j).value)
                    except:
                        pass
            except KeyError:
                pass
            lsd = LdStatData4(**data_line)
            sess.add(lsd)
    sess.commit()
    sess.close()


def load_death(sess):
    with open("death.txt", encoding='utf-8') as tsvfile:
        tsvreader = csv.reader(tsvfile, delimiter="\t")
        for i, line in enumerate(tsvreader):
            if i > 0:
                sgf = StatAgeDeath(age=line[0], prob=line[4].replace(',', '.').replace(' ', ''))
                sess.add(sgf)
        sess.commit()


if __name__ == '__main__':
    logging.basicConfig(level=logging.DEBUG, format='%(lineno)d %(asctime)s %(message)s')
    _sess = get_session()
    #xls_path = "data.xlsx"
    #process_file4(load_workbook(filename=xls_path), _sess)
    # cols1 = [c for c in char_range('E', 'Z')]
    #cols2 = ['A{0}'.format(c) for c in char_range('A', 'W')]
    # print(cols1 + cols2)
    #print(float('111 17,8'.replace(',', '.').replace(' ', '')))
    #print(float('11117.8'.replace(',', '.').replace(' ', '')))
    load_death(_sess)

